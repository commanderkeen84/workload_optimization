import time
import random
import requests
from openpyxl import load_workbook
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.formula.translate import Translator  # Import for formula translation

file_path = Path(r'.hotel_bookings.xlsx')  
worksheet_name = 'BR-cleanings'  # Define the worksheet name at the top

# region Defined Functions
def fetch_random_user():
    """
    Fetch a random user from the Random User API.
    Returns a tuple containing (name, email).
    """
    response = requests.get("https://randomuser.me/api/")
    if response.status_code == 200:
        data = response.json()
        first_name = data['results'][0]['name']['first']
        last_name = data['results'][0]['name']['last']
        email = data['results'][0]['email']
        name = f"{first_name} {last_name}"
        return name, email
    else:
        raise Exception("Failed to fetch random user data.")

def adjust_formulas(ws, start_row, num_rows):
    """
    Adjust formulas in the worksheet after inserting rows.
    :param ws: Worksheet object
    :param start_row: Row where new rows are inserted
    :param num_rows: Number of rows inserted
    """
    for row in ws.iter_rows(min_row=start_row + num_rows, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                # Translate the formula to adjust for the new rows
                cell.value = Translator(cell.value, origin=f"{cell.column_letter}{start_row}").translate_formula(f"{cell.column_letter}{start_row + num_rows}")

def calculate_br_cleaning_day(length_of_stay, check_in_date):
    """
    Calculate the BR-category based on the length of stay.
    :param length_of_stay: The length of stay category ("Mid", "Long", or "no BR").
    :param check_in_date: The check-in date as a datetime object.
    :return: The BR-cleaning day as a string.
    """
    if length_of_stay == "Mid":
        return (check_in_date + timedelta(days=7)).strftime('%d.%m.%Y')
    elif length_of_stay == "Long":
        return (check_in_date + timedelta(days=14)).strftime('%d.%m.%Y')
    else:
        return "no BR"

def generate_new_row(file_path, sheet_name, header_names):
    # Load the workbook and access the worksheet by name
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Find the column indices for the specified header names
    header_row = 2  # Assuming headers are in the second row
    column_indices = {}
    for col, cell in enumerate(ws[header_row], start=1):
        if cell.value in header_names:
            column_indices[cell.value] = col

    # Ensure all specified headers are found
    for header_name in header_names:
        if header_name not in column_indices:
            raise ValueError(f"Header '{header_name}' not found in the worksheet.")
    
    # Get the maximum reservation number from the "Reservation Number" column
    max_reservation_number = 0
    reservation_col = column_indices["Reservation Number"]
    for row in ws.iter_rows(min_row=3, min_col=reservation_col, max_col=reservation_col, values_only=True):
        if row[0] is not None and isinstance(row[0], int):
            max_reservation_number = max(max_reservation_number, row[0])

    # Generate a new reservation number that is higher than the maximum
    new_reservation_number = max_reservation_number + 1

    # Get existing room numbers from the "Room Number" column
    room_col = column_indices["Room Number"]
    existing_rooms = set()
    for row in ws.iter_rows(min_row=3, min_col=room_col, max_col=room_col, values_only=True):
        if row[0] is not None:
            existing_rooms.add(row[0])

    # Decide whether to add an individual user or a group booking
    is_group_booking = random.choice([True, False])  # Randomly decide

    if is_group_booking:
        # Generate group booking
        group_name = random.choice([
            "Harvard University","Stanford University","Massachusetts Institute of Technology (MIT)",
            "University of Cambridge", "University of Oxford","California Institute of Technology (Caltech)",
            "ETH Zurich – Swiss Federal Institute of Technology", "Imperial College London", "University of Chicago",
            "Yale University", "Princeton University", "Columbia University", "University of Pennsylvania",
            "University of California, Berkeley", "University of Edinburgh", "University of Michigan",
            "University of Copenhagen", "University of Amsterdam", "University of Zurich", "University of Helsinki",
            "University of Vienna", "University of Munich (LMU)", "University of Paris (Sorbonne)",
            "University of Bologna", "University of Barcelona", "University of Warsaw", "University of Stockholm",
            "University of Oslo", "University of Dublin (Trinity College)", "University of Geneva"
        ])
        
        group_size = random.randint(10, 20)  # Group size between 10 and 20
    
        # Generate Check-in and Check-out Dates
        check_in_date = datetime.now() + timedelta(days=random.randint(2, 365))
        check_out_date = check_in_date + timedelta(days=random.randint(14, 43))
        number_of_nights = (check_out_date - check_in_date).days
    	
        # Determine br-category 
        if 10 <= number_of_nights <= 14:
           br_category = "Mid"
        elif number_of_nights > 14:
            br_category = "Long"
        else:
            br_category = "no BR"
        
        # Calculate BR-cleaning day
        br_cleaning_day = calculate_br_cleaning_day(br_category, check_in_date)
        
        # Generate unique room numbers for the group
        group_data = []
        for _ in range(group_size):
            while True:
                room = random.randint(100, 1000)
                if room not in existing_rooms:
                    existing_rooms.add(room)
                    break
            group_data.append([
                new_reservation_number,
                room,
                group_name,
                f"{group_name.lower().replace(' ', '_')}@example.com",
                check_in_date.strftime('%d.%m.%Y'),
                check_out_date.strftime('%d.%m.%Y'),
                number_of_nights,
                br_category,
                br_cleaning_day
            ])
        wb.close()
        return group_data  # Return all rows for the group

    else:
        # Generate individual user booking
        while True:
            room = random.randint(100, 1000)
            if room not in existing_rooms:
                existing_rooms.add(room)
                break

        # Generate Check-in and Check-out Dates
        check_in_date = datetime.now() + timedelta(days=random.randint(2, 365))
        check_out_date = check_in_date + timedelta(days=random.randint(14, 43))
        number_of_nights = (check_out_date - check_in_date).days

        # Determine length of stay
        if 10 <= number_of_nights <= 14:
            br_category = "Mid"
        elif number_of_nights > 14:
            br_category = "Long"
        else:
            br_category = "no BR"
            
        # Calculate BR-cleaning day
        br_cleaning_day = calculate_br_cleaning_day(br_category, check_in_date)
       
        # Fetch random user data
        name, email = fetch_random_user()

        wb.close()  # Close the workbook after reading
        return [[
            new_reservation_number, 
            room, 
            name, 
            email, 
            check_in_date.strftime('%d.%m.%Y'),
            check_out_date.strftime('%d.%m.%Y'),
            number_of_nights,
            br_category, 
            br_cleaning_day
        ]]

def adjust_br_cleaning_day(target_day, cleaning_schedule, max_cleanings=5):
    """
    Adjust the BR-cleaning day to ensure no more than max_cleanings occur on the same day,
    while prioritizing days with zero cleanings and minimizing the deviation from the target day.
    :param target_day: The target BR-cleaning day (in DD.MM.YYYY format).
    :param cleaning_schedule: A dictionary tracking the number of cleanings per day.
    :param max_cleanings: The maximum number of cleanings allowed on the same day.
    :return: The adjusted BR-cleaning day.
    """
    target_date = datetime.strptime(target_day, '%d.%m.%Y')  # Convert to datetime
    best_date = None
    min_deviation = float('inf')  # Start with an infinitely large deviation

    # Step 1: Prioritize days with zero cleanings
    for days_shift in range(-7, 7):  # Adjust the range as needed
        adjusted_date = target_date + timedelta(days=days_shift)
        adjusted_str = adjusted_date.strftime('%d.%m.%Y')

        # Check if the date is completely free (0 BR-cleanings)
        if cleaning_schedule.get(adjusted_str, 0) == 0:
            cleaning_schedule[adjusted_str] = 1  # Mark the day as used
            return adjusted_str

    # Step 2: Minimize deviation if no free days are found
    for days_shift in range(-7, 7):
        adjusted_date = target_date + timedelta(days=days_shift)
        adjusted_str = adjusted_date.strftime('%d.%m.%Y')
        if cleaning_schedule.get(adjusted_str, 0) < max_cleanings:
            deviation = abs(days_shift)
            if deviation < min_deviation:
                best_date = adjusted_str
                min_deviation = deviation

    if best_date:
        cleaning_schedule[best_date] = cleaning_schedule.get(best_date, 0) + 1
        return best_date

    return target_day

def optimize_adjusted_days_simple(ws, target_col=9, adjusted_col=11, max_cleanings=5, iterations=3, search_window=14):
    """
    Iteratively improves the load-adjusted BR-cleaning day by re-checking each row.
    For each row, if a candidate day (within ±search_window days) is available (i.e., has fewer than
    max_cleanings assignments) and offers a lower absolute deviation from the target day,
    the candidate value is used to replace the current assignment.
    
    Parameters:
      ws           : The worksheet object.
      target_col   : Column number for Target BR-Cleaning Day (e.g., 9).
      adjusted_col : Column number for Load-adjusted BR-cleaning Day (e.g., 11).
      max_cleanings: Maximum allowed number of cleanings on the same day.
      iterations   : Number of passes over all rows.
      search_window: How many days (forward and back) to consider for alternative assignments.
    """
    for it in range(iterations):
        # Rebuild the cleaning schedule from the current assignments.
        cleaning_schedule = {}
        row_assignments = {}  # cache current assignments (row: day string) for ease of updating.
        for row in range(3, ws.max_row + 1):
            assigned = ws.cell(row=row, column=adjusted_col).value
            if assigned is not None:
                if isinstance(assigned, datetime):
                    assigned = assigned.strftime('%d.%m.%Y')
                row_assignments[row] = assigned
                cleaning_schedule[assigned] = cleaning_schedule.get(assigned, 0) + 1
        
        improvements = 0
        # Loop through each row and see if there's a better candidate.
        for row in range(3, ws.max_row + 1):
            # Retrieve target date from Column I.
            target_val = ws.cell(row=row, column=target_col).value
            if target_val is None:
                continue
            # Convert target to datetime if needed.
            if isinstance(target_val, datetime):
                target_date = target_val
            else:
                try:
                    target_date = datetime.strptime(target_val, '%d.%m.%Y')
                except ValueError:
                    continue  # skip rows with invalid target date format

            # Get current assignment.
            current_day_str = row_assignments.get(row)
            try:
                current_assigned = datetime.strptime(current_day_str, '%d.%m.%Y')
            except Exception:
                continue
            current_deviation = abs((current_assigned - target_date).days)
            
            best_candidate = current_day_str
            best_deviation = current_deviation

            # Search in the window for a candidate with lower deviation.
            for offset in range(-search_window, search_window + 1):
                candidate_date = target_date + timedelta(days=offset)
                candidate_str = candidate_date.strftime('%d.%m.%Y')
                # Only consider if candidate day hasn't reached the max limit.
                count = cleaning_schedule.get(candidate_str, 0)
                if candidate_str != current_day_str and count >= max_cleanings:
                    continue
                candidate_deviation = abs(offset)
                if candidate_deviation < best_deviation:
                    best_candidate = candidate_str
                    best_deviation = candidate_deviation

            # If a better candidate is found, update the cell and cleaning schedule.
            if best_candidate != current_day_str:
                # Decrement the count of the old candidate.
                cleaning_schedule[current_day_str] -= 1
                # Increment (or add) the count for the new candidate.
                cleaning_schedule[best_candidate] = cleaning_schedule.get(best_candidate, 0) + 1
                ws.cell(row=row, column=adjusted_col).value = best_candidate
                row_assignments[row] = best_candidate
                improvements += 1
        # Print iteration results (optional)
        print(f"Iteration {it+1}: {improvements} improvements made.")
        # If no improvements were made, exit early.
        if improvements == 0:
            break

# endregion

while True:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[worksheet_name]  # Use the worksheet name constant

    # Track cleaning schedule for initial adjustment
    cleaning_schedule = {}

    # Step 1: Process Existing Rows
    check_in_column = 5      # Column E (Check-in Date)
    check_out_column = 6     # Column F (Check-out Date)
    number_of_nights_column = 7  # Column G (Number of Nights)
    br_category_column = 8   # Column H (BR-Category)
    target_day_column = 9    # Column I (Target BR-Cleaning Day)
    adjusted_day_column = 11 # Column K (Adjusted BR-Cleaning Day)

    for row_idx in range(3, ws.max_row + 1):  # Start from row 3 (after the header)
        check_in_date = ws.cell(row=row_idx, column=check_in_column).value
        check_out_date = ws.cell(row=row_idx, column=check_out_column).value

        # Convert string dates to datetime objects if necessary
        if isinstance(check_in_date, str):
            try:
                check_in_date = datetime.strptime(check_in_date, '%d.%m.%Y')
            except ValueError:
                print(f"Row {row_idx}: Invalid Check-in Date format: {check_in_date}")
                continue

        if isinstance(check_out_date, str):
            try:
                check_out_date = datetime.strptime(check_out_date, '%d.%m.%Y')
            except ValueError:
                print(f"Row {row_idx}: Invalid Check-out Date format: {check_out_date}")
                continue

        if check_in_date is not None and check_out_date is not None:
            if isinstance(check_in_date, datetime) and isinstance(check_out_date, datetime):
                # Calculate the number of nights
                number_of_nights = (check_out_date - check_in_date).days
                ws.cell(row=row_idx, column=number_of_nights_column).value = number_of_nights

                # Determine BR-category
                if 10 <= number_of_nights <= 14:
                    br_category = "Mid"
                elif number_of_nights > 14:
                    br_category = "Long"
                else:
                    br_category = "no BR"
                ws.cell(row=row_idx, column=br_category_column).value = br_category

                # Calculate Target BR-Cleaning Day
                target_day = calculate_br_cleaning_day(br_category, check_in_date)
                ws.cell(row=row_idx, column=target_day_column).value = target_day

                # Initially assign Adjusted BR-Cleaning Day using the helper
                adjusted_day = adjust_br_cleaning_day(target_day, cleaning_schedule)
                ws.cell(row=row_idx, column=adjusted_day_column).value = adjusted_day
            else:
                print(f"Row {row_idx}: Invalid date format for Check-in or Check-out")
        else:
            print(f"Row {row_idx}: Missing Check-in or Check-out Date")

    # Populate cleaning_schedule with existing values in column K (Adjusted BR-Cleaning Day)
    adjusted_day_column = 11  # Column K
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=adjusted_day_column, max_col=adjusted_day_column, values_only=True):
        if row[0] is not None:
            if isinstance(row[0], datetime):
                formatted_date = row[0].strftime('%d.%m.%Y')
            else:
                formatted_date = row[0]
            cleaning_schedule[formatted_date] = cleaning_schedule.get(formatted_date, 0) + 1

    # --- Run the simplified iterative optimization step ---
    optimize_adjusted_days_simple(ws, target_col=target_day_column, adjusted_col=11, max_cleanings=5, iterations=3, search_window=7)

    # Generate and insert new data
    new_data = generate_new_row(file_path, worksheet_name, header_names=[
        "Room Number",
        "Reservation Number",
        "Target BR-Cleaning Day"
    ])

    for row_data in new_data:
        ws.insert_rows(3)  # Insert a new row at the top of the worksheet
        adjust_formulas(ws, start_row=3, num_rows=1)  # Adjust formulas after inserting each row

        for col, value in enumerate(row_data, start=1):
            ws.cell(row=3, column=col).value = value

        target_day = row_data[8]  # "Target BR-Cleaning Day" in column 9
        adjusted_day = adjust_br_cleaning_day(target_day, cleaning_schedule)
        ws.cell(row=3, column=11).value = adjusted_day

    # Update the COUNTIF formula for Column J
    br_cleaning_day_column = 9  # "Target BR-Cleaning Day" in column I
    formula_column = 10  # Column J
    for row in range(3, ws.max_row + 1):
        formula = f"=COUNTIF(I:I, I{row})"
        ws.cell(row=row, column=formula_column).value = formula

    # Update the COUNTIF formula for Column L (counting number of BR-cleanings)
    br_cleaning_day_column = 11  # "Adjusted BR-Cleaning Day" in column K
    formula_column = 12  # Column L
    for row in range(3, ws.max_row + 1):
        formula = f"=COUNTIF(K:K, K{row})"
        ws.cell(row=row, column=formula_column).value = formula

    # Update the days deviation from target in Column M
    formula_column = 13  # Column M
    for row in range(3, ws.max_row + 1):
        formula = f"=K{row}-I{row}"
        ws.cell(row=row, column=formula_column).value = formula

    wb.save(file_path)
    print(f"Added rows: {len(new_data)} and updated adjusted BR-cleaning days for all rows.")
    print(f"Added row at top: {new_data}")

    user_input = input("Press Enter to continue or type 'exit' to stop: ")
    if user_input.lower() == 'exit':
        print("Exiting the loop.")
        break

    # time.sleep(180)   # Uncomment to wait 3 minutes between loops